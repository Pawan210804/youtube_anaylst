# scripts/excel_report.py

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import BarChart, Reference
import os

# Load cleaned data
df = pd.read_csv("data/cleaned/youtube_clean.csv")

# Create workbook
wb = Workbook()

# ============ SHEET 1 — Raw Data ============
ws1 = wb.active
ws1.title = "YouTube Data"

# Header style
header_font    = Font(bold=True, color="FFFFFF", size=12)
header_fill    = PatternFill("solid", fgColor="1E3A5F")
header_align   = Alignment(horizontal="center", vertical="center")
border_side    = Side(style="thin", color="CCCCCC")
cell_border    = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)

# Write headers
headers = ["Video ID","Title","Channel","Category","Published At","Views","Likes","Comments","Engagement Rate","Hour","Day","Month"]
cols    = ["video_id","title","channel","category_id","published_at","views","likes","comments","engagement_rate","hour","day_of_week","month"]

for col_num, header in enumerate(headers, 1):
    cell = ws1.cell(row=1, column=col_num, value=header)
    cell.font      = header_font
    cell.fill      = header_fill
    cell.alignment = header_align
    cell.border    = cell_border

# Write data rows
for row_num, row in enumerate(df[cols].itertuples(index=False), 2):
    for col_num, value in enumerate(row, 1):
        cell = ws1.cell(row=row_num, column=col_num, value=value)
        cell.border    = cell_border
        cell.alignment = Alignment(horizontal="center")
        if row_num % 2 == 0:
            cell.fill = PatternFill("solid", fgColor="F0F4F8")

# Auto column width
for col in ws1.columns:
    max_len = max(len(str(cell.value or "")) for cell in col)
    ws1.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

# ============ SHEET 2 — Top 10 Videos ============
ws2 = wb.create_sheet("Top 10 Videos")

top10 = df.nlargest(10, "views")[["title","channel","views","likes","comments","engagement_rate"]]

top10_headers = ["Title","Channel","Views","Likes","Comments","Engagement Rate"]
for col_num, header in enumerate(top10_headers, 1):
    cell = ws2.cell(row=1, column=col_num, value=header)
    cell.font      = header_font
    cell.fill      = header_fill
    cell.alignment = header_align
    cell.border    = cell_border

for row_num, row in enumerate(top10.itertuples(index=False), 2):
    for col_num, value in enumerate(row, 1):
        cell = ws2.cell(row=row_num, column=col_num, value=value)
        cell.border    = cell_border
        cell.alignment = Alignment(horizontal="center")
        if row_num % 2 == 0:
            cell.fill = PatternFill("solid", fgColor="F0F4F8")

for col in ws2.columns:
    max_len = max(len(str(cell.value or "")) for cell in col)
    ws2.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

# ============ SHEET 3 — Summary Stats ============
ws3 = wb.create_sheet("Summary Stats")

stats = [
    ["Metric", "Value"],
    ["Total Videos",        len(df)],
    ["Total Views",         df["views"].sum()],
    ["Average Views",       int(df["views"].mean())],
    ["Average Likes",       int(df["likes"].mean())],
    ["Average Comments",    int(df["comments"].mean())],
    ["Avg Engagement Rate", f"{df['engagement_rate'].mean():.2f}%"],
    ["Most Active Day",     df.groupby("day_of_week")["engagement_rate"].mean().idxmax()],
    ["Best Hour to Post",   int(df.groupby("hour")["views"].mean().idxmax())],
]

for row_num, row in enumerate(stats, 1):
    for col_num, value in enumerate(row, 1):
        cell = ws3.cell(row=row_num, column=col_num, value=value)
        cell.border    = cell_border
        cell.alignment = Alignment(horizontal="center")
        if row_num == 1:
            cell.font = header_font
            cell.fill = header_fill
        elif row_num % 2 == 0:
            cell.fill = PatternFill("solid", fgColor="F0F4F8")

ws3.column_dimensions["A"].width = 25
ws3.column_dimensions["B"].width = 25

# Save
os.makedirs("output", exist_ok=True)
wb.save("output/youtube_report.xlsx")
print("✅ Excel report saved to output/youtube_report.xlsx")